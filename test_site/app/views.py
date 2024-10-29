from flask import Flask, redirect, render_template, url_for, request
from flask_simple_captcha import CAPTCHA
from . import app, db_svo_sequrity, mail
from .forms import NewCard, GetAppeal, GetSummary, AddNews, AddQuestion, SearchForm, Login_Svo_Sequrity, RegistrationForm
from .models import CardRazdel, News, User, get_title_post
import random
import re
from flask_mail import Message, Mail
import os
import os.path
from pathlib import Path
import xlrd
from werkzeug.utils import secure_filename
from flask_bcrypt import Bcrypt
from flask_login import UserMixin, LoginManager, logout_user, login_user, login_required, current_user
import openpyxl
from markupsafe import Markup
import sqlite3
BASEDIR=Path(__file__).parent
UPLOAD_FOLDER=BASEDIR/'static'/'images'
os.chdir('D:/OSPanel/domains/svo_test_site/test_site/my_content')
os.getcwd()
file_table_dop_education1='table_dop_education1.xlsx'
file_table_dop_education2='table_dop_education2.xlsx'
file_table_dop_education3='table_dop_education3.xlsx'
workbook=openpyxl.load_workbook('table_events_sequrity.xlsx')
worksheet=workbook.active

login_manager=LoginManager(app)
login_manager.login_view="index"
bcrypt=Bcrypt(app)
# def process_content(content):
#     # Регулярное выражение для поиска тегов <img>
#     img_pattern = re.compile(r'(<img[^>]+>)')

#     # Найти все изображения
#     images = img_pattern.findall(content)

#     # Удалить изображения из текста и заменить их временными маркерами
#     text_without_images = img_pattern.sub('{{img}}', content)

#     # Применить фильтр |safe к тексту без изображений
#     safe_text = text_without_images

#     # Восстановить изображения в тексте
#     for image in images:
#         safe_text = safe_text.replace('{{img}}', image, 1)

#     return safe_text
@login_manager.user_loader
def load_user(id):
    return User.query.get(int(id))

dict_events={}
for row in range(1, worksheet.max_row+1):
    key=worksheet.cell(row, 1).value
    value=worksheet.cell(row,2).value
    dict_events[key]=value
CAPTCHA_CONFIG = {
    'SECRET_CAPTCHA_KEY': 'WHITE RABBIT GO WITH YOU 13',
    'CAPTCHA_LENGTH': 6,
    'CAPTCHA_DIGITS': False,
    'EXPIRE_SECONDS': 600,
}
SIMPLE_CAPTCHA = CAPTCHA(config=CAPTCHA_CONFIG )
app = SIMPLE_CAPTCHA.init_app(app)
@app.route('/')
def index():
    title='Шереметьево безопасность. Главная страница'
    return render_template('index.html', title=title)
@app.route('/search', methods=['POST'])
def search():
    search_form=SearchForm()
    cards=CardRazdel.query
    if search_form.validate_on_submit():
        post_searched=search_form.searched.data
        cards=cards.filter(CardRazdel.content.like('%'+post_searched+'%'))
        cards=cards.order_by(CardRazdel.name_card).all()
        return render_template('search.html', title='Поиск по сайту', form=search_form, searched=post_searched, cards=cards)
@app.context_processor
def base():
    search_form=SearchForm()
    return dict(form=search_form)
@app.route('/about_company', methods=['GET'])
def about_company():
    return render_template('about_company.html', title='О компании')
@app.route('/education_center',  methods=['GET'])
def education_center():
    return render_template('education_center.html', title='Образовательный центр МАШ' )
@app.route('/applicants',  methods=['GET'])
def applicants():
    return render_template('applicants.html', title='Соискателям')
@app.route('/contacts',  methods=['GET'])
def contacts():
    return render_template('contacts.html', title='Контакты')
@app.route('/business_partners',  methods=['GET'])
def business_partners():
    return render_template('business_partners.html', title='Партнерам')
@app.route('/passenger',  methods=['GET'])
def passenger():
    cards_pass=CardRazdel.query.all()
    return render_template('passenger.html', title='Пассажирам', cards=cards_pass)

@app.route('/perevoz_objects',  methods=['GET'])
def perevoz_objects():
    return render_template('perevoz_objects.html', title='Перевоз оружия, спритных напитков и т.д.')
@app.route('/pass_pamytka',  methods=['GET'])
def pass_pamytka():
    return render_template('pass_pamytka.html', title='Памятка пассажира')
@app.route('/provoz_litiy',  methods=['GET'])
def provoz_litiy():
    return render_template('provoz_litiy.html', title='Провоз литиевых аккумуляторов')
@app.route('/special_passenger',  methods=['GET'])
def special_passenger():
    return render_template('special_passenger.html', title='Провоз литиевых аккумуляторов')
@app.route('/dosmotr_terminal',  methods=['GET'])
def dosmotr_terminal():
    return render_template('dosmotr_terminal.html', title='Досмотр в терминале')
@app.route('/talon',  methods=['GET'])
def talon():
    return render_template('talon.html', title='Электронный посалочный талон')

@app.route('/dont_put_off_shooes',  methods=['GET'])
def dont_put_off_shooes():
    return render_template('dont_put_off_shooes.html', title='Причины не снимать обувь и ремень (пояс)')
@app.route('/norma_pravo',  methods=['GET'])
def norma_pravo():
    return render_template('norma_pravo.html', title='Нормативная правовая база')
@app.route('/passengers_questions',  methods=['GET'])
def passengers_questions():
    return render_template('passengers_questions.html', title='Часто задаваемые вопросы')
@app.route('/forbidden_things',  methods=['GET'])
def forbidden_things():
    return render_template('forbidden_things.html', title='Запрещенные к провозу вещи')

@app.route('/storage_place',  methods=['GET'])
def storage_place():
    return render_template('storage_place.html', title='Хранение изъятых вещей')
@app.route('/hand_luggage',  methods=['GET'])
def hand_luggage():
    return render_template('hand_luggage.html', title='Ручная кладь')
@app.route('/sertificate_SanPin',  methods=['GET'])
def sertificate_SanPin():
    return render_template('sertificate_SanPin.html', title='Санитарные нормы оборудования')

@app.route('/add_question',  methods=['GET', 'POST'])
def add_question():
    form=AddQuestion()
    new_captcha_dict = SIMPLE_CAPTCHA.create()
    if  request.method=='POST':
        c_hash = request.form.get('captcha-hash')
        c_text = request.form.get('captcha-text')
        if SIMPLE_CAPTCHA.verify(c_text, c_hash):
            name=form.fio.data
            text_message=form.text_letter.data
            email=form.email_visitor.data
            print(name, email, text_message)
            msg = Message(
                'Вопросы касающиеся перевозки грузов',
                sender=email,
                recipients=['juliatalaeva8881@gmail.com'],
                body=text_message+'\n'+f'ФИО адресата: {name}\nE-Mail: {email}\n\n{text_message}'
            )
            mail.send(msg)
            return redirect(url_for('index'))
        else:
            return 'Вы не прошли проверку на робота'
    return render_template('add_question.html', title='Задать вопрос', form=form, captcha=new_captcha_dict)
@app.route('/add_summary',  methods=['GET', 'POST'])
def add_summary():
    form=GetSummary()
    ddrug=''
    aaccept=''
    ppasport=''
    new_captcha_dict = SIMPLE_CAPTCHA.create()
    if  form.validate_on_submit():
        c_hash = request.form.get('captcha-hash')
        c_text = request.form.get('captcha-text')
        if SIMPLE_CAPTCHA.verify(c_text, c_hash):
            lastname=form.lastname.data
            name=form.name.data
            second_name=form.second_name.data
            phone=form.phone_number.data
            education=form.education.data
            pasport=form.have_Russian_pasport.data
            drugs=form.dont_use_drugs.data
            accept=form.accept_prov.data
            print(lastname, name, second_name, phone, pasport, drugs, accept)
            if (drugs==True):
                ddrug='да'
            if (accept==True):
                aaccept='да'
            if (pasport==True):
                ppasport='да'
            msg = Message(
                'Отклик соискателя',
                sender='juliatalaeva8881@gmail.com',
                recipients=['juliatalaeva8881@gmail.com'],
                body=f'ФИО адресата:{lastname} {name} {second_name}\nНомер телефона: {phone}\nУровень образования: {education}\nИмеет гражданство РФ: {ppasport}\nОтсутствует судимость: {ddrug}\nСогласие на обработку персональных данных: {aaccept}'
            )
            mail.send(msg)
            return redirect(url_for('index'))
        else:
            return 'Вы не прошли проверку на робота'
    return render_template('add_summary.html', title='Подать заявление на работу', form=form, captcha=new_captcha_dict)

@app.route('/add_letter',  methods=['GET', 'POST'])
def add_letter():
    form=GetAppeal()
    new_captcha_dict = SIMPLE_CAPTCHA.create()
    if  request.method=='POST':
        c_hash = request.form.get('captcha-hash')
        c_text = request.form.get('captcha-text')
        if SIMPLE_CAPTCHA.verify(c_text, c_hash):
            name=form.fio.data
            text_message=form.text_letter.data
            email=form.email_visitor.data
            print(name, email, text_message)
            msg = Message(
                'Обращение по вопросам работы организации',
                sender=email,
                recipients=['juliatalaeva8881@gmail.com'],
                body=text_message+'\n'+f'ФИО адресата: {name}\nE-Mail: {email}\n\n{text_message}'
            )
            mail.send(msg)
            return redirect(url_for('index'))
        else:
            return 'Вы не прошли проверку на робота'
    return render_template('add_letter.html', form=form, captcha=new_captcha_dict)
   
@app.route('/type_appeal',  methods=['GET', 'POST'])
def type_appeal():
    return render_template('type_appeal.html', title='Тип обращения')
@app.route('/education_statement',  methods=['GET', 'POST'])
def education_statement():
    return render_template('education_statement.html/', title='Заявка на обучение')
@app.route('/add_card',  methods=['GET', 'POST'])
@login_required
def add_card():
    form=NewCard()
    if form.validate_on_submit():
        card=CardRazdel()
        card.name_card=form.name.data
        card.content=form.content.data
        card.address=form.adress.data
        image=form.image.data
        image_name=secure_filename(image.filename)
        UPLOAD_FOLDER.mkdir(exist_ok=True)
        image.save(UPLOAD_FOLDER/image_name)
        card.image=image_name
        db_svo_sequrity.session.add(card)
        db_svo_sequrity.session.commit()
        return redirect(url_for('add_card'))
    return render_template('add_card.html', title='Добавить новый раздел', form=form)
@app.route('/responsive_pass',  methods=['GET', 'POST'])
def responsive_pass():
    return render_template('responsive_pass.html', title='Ответствтенность пассажира')
@app.route('/news',  methods=['GET', 'POST'])
def news():
    news=News.query.all()
    return render_template('news.html', title='Новости', news=news)
@app.route('/news_more/<alias>',  methods=['GET', 'POST'])
def news_more(alias):
    title, post=get_title_post(alias)
    t=Markup(title)
    p=Markup(post)
    return render_template('news_more.html', title=t, post=p)
@app.route('/add_news',  methods=['GET', 'POST'])
@login_required
def add_news():
    form=AddNews()
    if request.method=='POST':
        news=News()
        news.news_name=form.name_news.data
        # text=form.text_news.data
        # processed_content = process_content(text)
        # news.text=processed_content 
        news.text=form.text_news.data
        news.url=form.url.data
        news.created_date=form.date_news.data
        db_svo_sequrity.session.add(news)
        db_svo_sequrity.session.commit()
        return redirect(url_for('news_more', alias=form.url.data))
    return render_template('add_news.html', title='Добавить новость', form=form)
@app.route('/sertificates_and_linc',  methods=['GET', 'POST'])
def sertificates_and_linc():
    return render_template('sertificates_and_linc.html', title='Сертификаты и лицензии')
@app.route('/hunters',  methods=['GET', 'POST'])
def hunters():
    return render_template('hunters.html', title='Охотникам')
@app.route('/position_on_market',  methods=['GET', 'POST'])
def position_on_market():
    return render_template('position_on_market.html', title='Позиция на рынке')
@app.route('/inter_managment_system',  methods=['GET', 'POST'])
def inter_managment_system():
    return render_template('inter_managment_system.html', title='Интегрированная система менеджмента АО "Щереметьево Безопасность"')
@app.route('/personal_organization',  methods=['GET', 'POST'])
def personal_organization():
    return render_template('personal_organization.html', title='Персонал организации')
@app.route('/info_about_sells_and_trafics',  methods=['GET', 'POST'])
def info_about_sells_and_trafics():
    return render_template('info_about_sells_and_trafics.html', title='Тарифы и цены на оказание услуг')
@app.route('/education_video',  methods=['GET', 'POST'])
def education_video():
    return render_template('education_video.html', title='Учебное видео')
@app.route('/other_info',  methods=['GET', 'POST'])
def other_info():
    return render_template('other_info.html', title='Прочее')
@app.route('/about_educational_center',  methods=['GET', 'POST'])
def about_educational_center():
    return render_template('about_educational_center.html/', title='Сведения об учебном центре')
@app.route('/linc_sertif_edu',  methods=['GET', 'POST'])
def linc_sertif_edu():
    return render_template('linc_sertif_edu.html/', title='Лицензии и сертификаты')
@app.route('/defend_job',  methods=['GET', 'POST'])
def defend_job():
    return render_template('defend_job.html', title='Охрана труда', events=dict_events)
@app.route('/plan_education',  methods=['GET', 'POST'])
def plan_education():
    return render_template('plan_education.html/', title='План обучения')
@app.route('/edu_center_annotations',  methods=['GET', 'POST'])
def edu_center_annotations():
    return render_template('edu_center_annotations.html/', title='Учебный центр--аннотации')
@app.route('/program_dop_education',  methods=['GET', 'POST'])
def program_dop_education():
    book1 = openpyxl.load_workbook(file_table_dop_education1)
    sheet1 = book1.active
    book2 = openpyxl.load_workbook(file_table_dop_education2)
    sheet2= book2.active
    book3 = openpyxl.load_workbook(file_table_dop_education3)
    sheet3 = book3.active
    return render_template('program_dop_education.html', title='Программы дополнительного профессионального образования', sheet1=sheet1, sheet2=sheet2, sheet3=sheet3)
@app.route('/available_environment',  methods=['GET', 'POST'])
def available_environment():
    return render_template('available_environment.html', title='Доступная среда')
@app.route('/international_cooperation',  methods=['GET', 'POST'])
def international_cooperation():
    return render_template('international_cooperation.html', title='Международное сотрудничество')
@app.route('/vacants_places',  methods=['GET', 'POST'])
def vacants_places():
    return render_template('vacants_places.html', title='Вакантные места')
@app.route('/the_financial_part',  methods=['GET', 'POST'])
def the_financial_part():
    return render_template('the_financial_part.html', title='Финансовой-хозяйственная деятельность')
@app.route('/mat_teq_provision',  methods=['GET', 'POST'])
def mat_teq_provision():
    return render_template('mat_teq_provision.html', title='Материально-техничсекое обеспечение')
@app.route('/educ_doc',  methods=['GET', 'POST'])
def educ_doc():
    return render_template('educ_doc.html', title='Документы')
@app.route('/edu_education',  methods=['GET', 'POST'])
def edu_education():
    return render_template('edu_education.html', title='Образование')
@app.route('/edu_structure',  methods=['GET', 'POST'])
def edu_structure():
    return render_template('edu_structure.html', title='Структура управления')
@app.route('/info_edu_center',  methods=['GET', 'POST'])
def info_edu_center():
    return render_template('info_edu_center.html', title='Основная информация')
@app.route('/edu_pedagog',  methods=['GET', 'POST'])
def edu_pedagog():
    return render_template('edu_pedagog.html', title='Преподавательский состав')
@app.route('/education_service',  methods=['GET', 'POST'])
def education_service():
    return render_template('education_service.html', title='Платные образовательные услуги')
@app.route('/admin_page',  methods=['GET', 'POST'])
@login_required
def admin_page():
    return render_template('admin_page.html', title='Настройки и администрирование')
@app.route('/register',  methods=['GET', 'POST'])
@login_required
def register():
    form=RegistrationForm()
    if request.method=='POST':
        hashed_password=bcrypt.generate_password_hash(form.password.data)
        new_user=User(username=form.username.data, password=hashed_password, fio=form.fio.data)
        db_svo_sequrity.session.add(new_user)
        db_svo_sequrity.session.commit()
        return redirect(url_for('log_in_admin_svo_sequrity'))
    return render_template('register.html', title='Регистрация админа', form=form)

@app.route('/log_in_admin_svo_sequrity',  methods=['GET', 'POST'])
def log_in_admin_svo_sequrity():
    form=Login_Svo_Sequrity()
    if request.method=='POST':
        users=User.query
        user= users.filter_by(username=form.username.data).first()
        if user:
            if bcrypt.check_password_hash(user.password, form.password.data):
                login_user(user)
                return redirect('admin_page')
        else:
            return redirect('index')
    return render_template('log_in_admin_svo_sequrity.html', title='Вход в админ панель', form=form)
@app.route('/logout/')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404
@app.errorhandler(500)
def page_not_found(e):
    return render_template('500.html'), 500
@app.errorhandler(502)
def page_not_found(e):
    return render_template('502.html'), 502
db_svo_sequrity.create_all()
if __name__=='__main__':
    app.run(debug=True)
    user=db.filter_by(username=form.username.data).first()